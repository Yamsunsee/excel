import math
import random
import openpyxl

MAIN_PATH = "C:\\Users\\Admin\\Desktop\\Python"
IMPORT_FILE_NAME = "data"
EXPORT_FILE_NAME = "result"


def shuffle(data):
    for item in data:
        random.shuffle(item["answers"])
    return random.shuffle(data)


def exportFile(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    length = len(data)

    for i in range(length):
        for j in range(5):
            if not j:
                question = data[i]["question"]
                questionContent = question.split(".")
                questionContent = ".".join(questionContent[1:])
                sheet.cell(row=i * 5 + j + 1, column=1).value = f"Câu {i+1}:{questionContent}"
            else:
                answer = data[i]["answers"][j - 1]
                units = "a b c d".split()
                sheet.cell(row=i * 5 + j + 1, column=1).value = f"{units[j - 1]}.{answer[2:]}"

    workbook.save(MAIN_PATH + f"\\{EXPORT_FILE_NAME}.xlsx")


def main(path):
    # Import file
    workbook = openpyxl.load_workbook(path + f"\\{IMPORT_FILE_NAME}.xlsx")
    sheet = workbook.active
    max_row = sheet.max_row

    result = []
    for i in range(math.ceil(max_row // 5)):
        row = {}
        row["answers"] = []

        for j in range(5):
            if not j:
                row["question"] = sheet.cell(row=i * 5 + j + 1, column=1).value
            else:
                row["answers"] += [sheet.cell(row=i * 5 + j + 1, column=1).value]

        result += [row]

    # Shuffle file
    shuffle(result)

    # # Export file
    exportFile(result)

    return "Done!"


if __name__ == "__main__":
    state = main(MAIN_PATH)
    print(state)